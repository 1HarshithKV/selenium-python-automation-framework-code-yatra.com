import inspect
import logging
import softest
from openpyxl import workbook, load_workbook
import csv

class Utils(softest.TestCase):
    def assertListItemText(self,list,values): #1 Stop
        for stop in list:
            print("The text is: " + stop.text[0:6])
            self.soft_assert(self.assertEqual, stop.text[0:6],values)
            # stop.text == '1 Stop'
            # print("Assert Pass")
            if stop.text[0:6] == values:
                print("Test Passed")
            else:
                print("Test Failed")
        self.assert_all()

    def custom_logger(loglevel=logging.DEBUG):
        #set class/method name from where it is called
        logger_name = inspect.stack()[1][3]
        #create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(loglevel)
        #create file/console handler and set log level
        fh = logging.FileHandler("report/automation.log", mode='a') #by default it is write mode
        #create formatter - how you want your logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s - %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
        #add formatter to console or file handler
        fh.setFormatter(formatter)
        #add file/console handler to logger
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name,sheet):
        data_list = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct+1):
            row = []
            for j in range(1, col_ct+1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)
        return data_list

    def read_data_from_csv(file_name):
        data_list = [] #create an empty list
        csvdata = open(file_name,"r") #open csv file in read mode
        reader = csv.reader(csvdata) #create csv reader
        next(reader) #skip header
        for rows in reader:         #add csv rows to list
            data_list.append(rows)
        return data_list







